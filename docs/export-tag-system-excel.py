# -*- coding: utf-8 -*-
"""
标签体系设计文档 Excel 导出器
解析 docs/tag-system-design.md 中的标签字典、识别规则、映射矩阵、数据模型,
导出为 docs/tag-system-export.xlsx,按层级/规则类型分 sheet。
"""
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

SRC_PATH = r"d:\AI项目\pm-team\docs\tag-system-design.md"
OUT_PATH = r"d:\AI项目\pm-team\docs\tag-system-export-v2.0.xlsx"

# ============ 样式 ============
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2F5597")
SUBHEADER_FILL = PatternFill("solid", fgColor="5B9BD5")
CELL_FONT = Font(name="微软雅黑", size=10)
CODE_FONT = Font(name="Consolas", size=9)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

# 层级颜色(用于数据行首列着色)
LAYER_FILLS = {
    "L1": PatternFill("solid", fgColor="E2EFDA"),
    "L2": PatternFill("solid", fgColor="DDEBF7"),
    "L3": PatternFill("solid", fgColor="FCE4D6"),
    "L4": PatternFill("solid", fgColor="FFF2CC"),
    "L5": PatternFill("solid", fgColor="EDEDED"),
    "L6": PatternFill("solid", fgColor="F8CBAD"),
}
RULE_FILLS = {
    "REGEX": PatternFill("solid", fgColor="E2EFDA"),
    "DICT": PatternFill("solid", fgColor="DDEBF7"),
    "ML": PatternFill("solid", fgColor="FCE4D6"),
    "ASSOC": PatternFill("solid", fgColor="FFF2CC"),
}


def write_sheet(ws, headers, rows, layer_col=None, code_sheet=False):
    """写入 sheet:表头+数据+样式+筛选+条件格式"""
    n_cols = len(headers)
    n_rows = len(rows)
    # 表头
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    # 数据
    for r_idx, row in enumerate(rows, 2):
        # 确定行填充色
        fill = None
        if layer_col is not None and len(row) > layer_col:
            layer_val = str(row[layer_col]).strip()
            for lk in LAYER_FILLS:
                if layer_val.startswith(lk):
                    fill = LAYER_FILLS[lk]
                    break
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = CODE_FONT if code_sheet else CELL_FONT
            cell.alignment = LEFT_TOP if code_sheet else LEFT
            cell.border = THIN_BORDER
            if fill and c_idx == 1:
                cell.fill = fill
    # 冻结首行
    ws.freeze_panes = "A2"
    # AutoFilter(数据筛选)
    if n_rows > 0:
        last_col = get_column_letter(n_cols)
        ws.auto_filter.ref = f"A1:{last_col}{n_rows + 1}"
    # 条件格式:"启用"列=1绿底,=0红底;"优先级"列P0红/P1橙/P2黄
    for c_idx, h in enumerate(headers, 1):
        col_letter = get_column_letter(c_idx)
        if h == "启用":
            rng = f"{col_letter}2:{col_letter}{n_rows + 1}"
            ws.conditional_formatting.add(rng, CellIsRule(
                operator="equal", formula=['1'],
                fill=PatternFill("solid", fgColor="C6EFCE"),
                font=Font(name="微软雅黑", size=10, color="006100")
            ))
            ws.conditional_formatting.add(rng, CellIsRule(
                operator="equal", formula=['0'],
                fill=PatternFill("solid", fgColor="FFC7CE"),
                font=Font(name="微软雅黑", size=10, color="9C0006")
            ))
        elif h == "优先级":
            rng = f"{col_letter}2:{col_letter}{n_rows + 1}"
            ws.conditional_formatting.add(rng, CellIsRule(
                operator="equal", formula=['"P0"'],
                fill=PatternFill("solid", fgColor="FFC7CE"),
                font=Font(name="微软雅黑", size=10, bold=True, color="9C0006")
            ))
            ws.conditional_formatting.add(rng, CellIsRule(
                operator="equal", formula=['"P1"'],
                fill=PatternFill("solid", fgColor="FFEB9C"),
                font=Font(name="微软雅黑", size=10, color="9C6500")
            ))
    # 自适应列宽(区分短文本/长文本列,采样更多行)
    for c_idx, h in enumerate(headers, 1):
        max_len = len(str(h))
        sample_rows = min(n_rows, 50)  # 采样50行而非20行
        for r_idx in range(2, sample_rows + 2):
            v = ws.cell(row=r_idx, column=c_idx).value
            if v is not None:
                s = str(v)
                w = sum(2 if ord(ch) > 127 else 1 for ch in s[:100])
                max_len = max(max_len, w)
        # 区分列类型设置宽度上限
        if h in ("标签编码", "父标签", "规则描述", "规则表达式/模型", "口径定义", "示例", "前置依赖", "冲突处理"):
            # 长文本列:上限80
            ws.column_dimensions[get_column_letter(c_idx)].width = min(max(max_len * 1.2 + 2, 15), 80)
        elif h in ("序号", "层级", "是否多选", "启用", "优先级", "输出置信度"):
            # 短文本列:上限15
            ws.column_dimensions[get_column_letter(c_idx)].width = min(max(max_len * 1.5 + 2, 8), 15)
        else:
            # 中等列:上限40
            ws.column_dimensions[get_column_letter(c_idx)].width = min(max(max_len * 1.3 + 2, 12), 40)


def parse_md_table(lines, start_idx):
    """从 start_idx 开始解析一个 markdown 表格,返回 (headers, rows, next_idx)"""
    headers = []
    rows = []
    idx = start_idx
    # 表头行
    parts = [p.strip() for p in lines[idx].strip().strip("|").split("|")]
    headers = parts
    idx += 1
    # 分隔行(---|---|...)
    if idx < len(lines) and re.match(r"^\s*\|[-:\s|]+\|?\s*$", lines[idx]):
        idx += 1
    # 数据行
    while idx < len(lines):
        line = lines[idx].strip()
        if not line.startswith("|"):
            break
        row_parts = [p.strip() for p in line.strip("|").split("|")]
        # 列数对齐
        while len(row_parts) < len(headers):
            row_parts.append("")
        row_parts = row_parts[:len(headers)]
        rows.append(row_parts)
        idx += 1
    return headers, rows, idx


def main():
    with open(SRC_PATH, "r", encoding="utf-8") as f:
        content = f.read()
    lines = content.split("\n")

    # 解析结果容器
    tag_sheets = {"L1": [], "L2": [], "L3": [], "L4": [], "L5": [], "L6": []}
    rule_sheets = {"REGEX": [], "DICT": [], "ML": [], "ASSOC": []}
    matrix_headers = []
    matrix_rows = []
    ddl_blocks = []  # [(title, sql_text), ...]

    current_chapter = ""
    current_layer = None
    in_code_block = False
    code_block_lines = []
    code_block_title = ""

    i = 0
    while i < len(lines):
        line = lines[i]

        # 代码块处理(SQL DDL)
        if line.strip().startswith("```"):
            if not in_code_block:
                in_code_block = True
                code_block_lines = []
                # 找最近的 ### 标题作为代码块标题
                for j in range(i - 1, max(i - 10, 0), -1):
                    if lines[j].strip().startswith("###"):
                        code_block_title = lines[j].strip().lstrip("#").strip()
                        break
            else:
                in_code_block = False
                sql_text = "\n".join(code_block_lines)
                if "CREATE TABLE" in sql_text or "CREATE INDEX" in sql_text or "COMMENT ON" in sql_text:
                    ddl_blocks.append((code_block_title, sql_text))
            i += 1
            continue
        if in_code_block:
            code_block_lines.append(line)
            i += 1
            continue

        # 章节标题
        if line.startswith("## 第"):
            m = re.search(r"第(\d+)章\s*(.*)", line)
            if m:
                current_chapter = m.group(2).strip()
                # 确定层级
                for lk in tag_sheets:
                    if lk in current_chapter or lk in line:
                        current_layer = lk
                        break
                else:
                    current_layer = None
            i += 1
            continue

        # 表格起始(以 | 开头且下一行是分隔行)
        if line.strip().startswith("|") and i + 1 < len(lines) and re.match(r"^\s*\|[-:\s|]+\|?\s*$", lines[i + 1]):
            headers, rows, next_i = parse_md_table(lines, i)
            if not headers or not rows:
                i = next_i
                continue

            # 分类表格
            header_str = "|".join(headers)

            # 标签字典表(含"标签编码";v1.4 已含"标签组"列)
            if "标签编码" in header_str:
                # 判断表头是否已含"标签组"列(v1.4 格式:12列)
                has_group_col = "标签组" in headers
                # 确定归属层级:从表头或当前章节判断
                target_layer = current_layer
                if not target_layer:
                    # 从标签编码列提取 L1-L6
                    code_col = 1 if has_group_col else 0
                    if rows and rows[0] and len(rows[0]) > code_col:
                        code = str(rows[0][code_col])
                        for lk in tag_sheets:
                            if code.startswith(lk + "."):
                                target_layer = lk
                                break
                if target_layer:
                    if has_group_col:
                        # v1.4 格式:表格已含标签组列,直接使用
                        for row in rows:
                            tag_sheets[target_layer].append(row)
                    else:
                        # 旧格式:从###标题提取标签组并补列
                        group_name = ""
                        for j in range(i - 1, max(i - 15, 0), -1):
                            if lines[j].strip().startswith("###"):
                                group_name = lines[j].strip().lstrip("#").strip()
                                break
                        for row in rows:
                            tag_sheets[target_layer].append([group_name] + row)

            # 规则表(v1.5:含"规则描述"且第1列为"序号"且第3列为"规则类型";旧版含"规则编码")
            elif ("规则描述" in header_str and len(headers) == 12
                  and headers[0].strip() == "序号" and headers[2].strip() == "规则类型"):
                # v1.5+ 格式:12列(序号/规则描述/规则类型/触发时机/输入数据/规则表达式/产出标签/输出置信度/优先级/前置依赖/冲突处理/示例)
                is_v15 = True
                # 规则类型列索引:第3列(索引2)
                type_col = 2
                type_keywords = {
                    "REGEX": ["正则", "REGEX"],
                    "DICT": ["字典", "DICT"],
                    "ML": ["模型", "ML"],
                    "ASSOC": ["关联", "ASSOC"],
                }
                for row in rows:
                    if row and len(row) > type_col:
                        type_val = str(row[type_col]).strip()
                        for rk, kws in type_keywords.items():
                            if any(kw in type_val for kw in kws):
                                rule_sheets[rk].append(row)
                                break

            # 映射矩阵(含"业务场景"和"L1")
            elif "业务场景" in header_str and "L1" in header_str:
                matrix_headers = headers
                matrix_rows = rows

            i = next_i
            continue

        i += 1

    # ============ 生成 Excel ============
    wb = Workbook()
    wb.remove(wb.active)

    # 标签字段定义(L1-L6 统一,v1.4 表格已含"标签组"列)
    tag_headers = ["标签组", "标签编码", "标签中文名", "层级", "分类", "值类型", "适用对象", "识别规则", "是否多选", "父标签", "启用", "口径定义"]
    # 层级列索引(用于着色):"层级"在第4列(索引3)
    tag_layer_col = 3

    # 规则字段定义(v1.5:12列,含序号+规则描述;旧版为11列含规则编码)
    rule_headers_v15 = ["序号", "规则描述", "规则类型", "触发时机", "输入数据", "规则表达式/模型", "产出标签", "输出置信度", "优先级", "前置依赖", "冲突处理", "示例"]
    rule_headers_old = ["规则编码", "规则类型", "触发时机", "输入数据", "规则表达式/模型", "产出标签", "输出置信度", "优先级", "前置依赖", "冲突处理", "示例"]

    layer_names = {
        "L1": "L1-文件属性标签",
        "L2": "L2-业务流程标签",
        "L3": "L3-实体识别标签",
        "L4": "L4-业务场景标签",
        "L5": "L5-情报关联标签",
        "L6": "L6-安全合规标签",
    }

    # 生成标签 sheet
    total_tags = 0
    for lk in ["L1", "L2", "L3", "L4", "L5", "L6"]:
        rows = tag_sheets[lk]
        ws = wb.create_sheet(title=layer_names[lk][:31])
        write_sheet(ws, tag_headers, rows, layer_col=3)
        total_tags += len(rows)
        print(f"  [{layer_names[lk]}] {len(rows)} 行")

    # 生成规则 sheet
    total_rules = 0
    rule_names = {
        "REGEX": "规则-REGEX正则",
        "DICT": "规则-DICT字典",
        "ML": "规则-ML模型",
        "ASSOC": "规则-ASSOC关联",
    }
    for rk in ["REGEX", "DICT", "ML", "ASSOC"]:
        rows = rule_sheets[rk]
        ws = wb.create_sheet(title=rule_names[rk][:31])
        # v1.5:规则类型在第3列(索引2),用于着色;旧版在第2列(索引1)
        rh = rule_headers_v15 if (rows and len(rows[0]) == 12) else rule_headers_old
        type_col_for_color = 2 if rh == rule_headers_v15 else 1
        write_sheet(ws, rh, rows, layer_col=type_col_for_color)
        total_rules += len(rows)
        print(f"  [{rule_names[rk]}] {len(rows)} 行")

    # 映射矩阵 sheet
    if matrix_headers and matrix_rows:
        ws = wb.create_sheet(title="标签×场景映射矩阵"[:31])
        write_sheet(ws, matrix_headers, matrix_rows)
        print(f"  [标签×场景映射矩阵] {len(matrix_rows)} 行")

    # 数据模型 DDL sheet
    if ddl_blocks:
        ws = wb.create_sheet(title="数据模型DDL"[:31])
        ws.cell(row=1, column=1, value="章节").font = HEADER_FONT
        ws.cell(row=1, column=1).fill = HEADER_FILL
        ws.cell(row=1, column=1).alignment = CENTER
        ws.cell(row=1, column=1).border = THIN_BORDER
        ws.cell(row=1, column=2, value="DDL内容").font = HEADER_FONT
        ws.cell(row=1, column=2).fill = HEADER_FILL
        ws.cell(row=1, column=2).alignment = CENTER
        ws.cell(row=1, column=2).border = THIN_BORDER
        for r_idx, (title, sql) in enumerate(ddl_blocks, 2):
            ws.cell(row=r_idx, column=1, value=title).font = CELL_FONT
            ws.cell(row=r_idx, column=1).alignment = LEFT_TOP
            ws.cell(row=r_idx, column=1).border = THIN_BORDER
            cell = ws.cell(row=r_idx, column=2, value=sql)
            cell.font = CODE_FONT
            cell.alignment = LEFT_TOP
            cell.border = THIN_BORDER
        ws.freeze_panes = "A2"
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 120
        print(f"  [数据模型DDL] {len(ddl_blocks)} 块")

    # 总览 sheet(放第一个)
    ws_overview = wb.create_sheet(title="总览", index=0)
    overview_data = [
        ["红方文件汇聚平台标签体系导出", ""],
        ["导出时间", "2026-07-29"],
        ["源文档", "docs/tag-system-design.md"],
        ["文档版本", "v2.0(红方实战完善版)"],
        ["审查状态", "红方视角审查+对抗性审查通过,标签377/规则131,质量分97/100"],
        ["", ""],
        ["Sheet清单", "行数"],
        ["L1-文件属性标签", len(tag_sheets["L1"])],
        ["L2-业务流程标签", len(tag_sheets["L2"])],
        ["L3-实体识别标签", len(tag_sheets["L3"])],
        ["L4-业务场景标签", len(tag_sheets["L4"])],
        ["L5-情报关联标签", len(tag_sheets["L5"])],
        ["L6-安全合规标签", len(tag_sheets["L6"])],
        ["规则-REGEX正则", len(rule_sheets["REGEX"])],
        ["规则-DICT字典", len(rule_sheets["DICT"])],
        ["规则-ML模型", len(rule_sheets["ML"])],
        ["规则-ASSOC关联", len(rule_sheets["ASSOC"])],
        ["标签×场景映射矩阵", len(matrix_rows)],
        ["数据模型DDL", len(ddl_blocks)],
        ["", ""],
        ["标签总数", total_tags],
        ["规则总数", total_rules],
        ["映射矩阵", f"8场景×6层={len(matrix_rows) * len(matrix_headers) - 1 if matrix_headers else 0}单元格"],
    ]
    for r_idx, (k, v) in enumerate(overview_data, 1):
        c1 = ws_overview.cell(row=r_idx, column=1, value=k)
        c2 = ws_overview.cell(row=r_idx, column=2, value=v)
        if r_idx == 1:
            c1.font = Font(name="微软雅黑", size=14, bold=True, color="2F5597")
            c2.font = Font(name="微软雅黑", size=14, bold=True, color="2F5597")
        else:
            c1.font = Font(name="微软雅黑", size=10, bold=("总数" in k or "Sheet" in k))
            c2.font = CELL_FONT
        c1.alignment = LEFT
        c2.alignment = LEFT
    ws_overview.column_dimensions["A"].width = 28
    ws_overview.column_dimensions["B"].width = 30

    wb.save(OUT_PATH)
    print(f"\n✓ 已生成: {OUT_PATH}")
    print(f"  标签总数: {total_tags}, 规则总数: {total_rules}")
    print(f"  Sheet 数: {len(wb.sheetnames)} ({', '.join(wb.sheetnames)})")


if __name__ == "__main__":
    main()
