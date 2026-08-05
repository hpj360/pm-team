# -*- coding: utf-8 -*-
"""
监控指标体系 Mock 数据 Excel 生成器
基于 docs/monitor-design.md 数据模型 DDL 生成,每个 sheet 100 行数据
输出: docs/monitor-metrics-mock.xlsx
"""
import random
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

random.seed(20260727)  # 固定种子,保证可复现

OUT_PATH = r"d:\AI项目\pm-team\docs\monitor-metrics-mock.xlsx"
ROWS_PER_SHEET = 100

# ============ 样式 ============
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2F5597")
CELL_FONT = Font(name="微软雅黑", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def write_sheet(ws, headers, rows):
    """写入一个 sheet: 表头加粗+底色,数据居中,冻结首行,自适应列宽"""
    # 表头
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    # 数据
    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = CELL_FONT
            cell.alignment = LEFT
            cell.border = THIN_BORDER
    # 冻结首行
    ws.freeze_panes = "A2"
    # 自适应列宽(按表头与样本长度估算)
    for c_idx, h in enumerate(headers, 1):
        max_len = len(str(h))
        for r_idx in range(2, min(len(rows) + 2, 12)):  # 抽样前10行
            v = ws.cell(row=r_idx, column=c_idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max(max_len * 1.8 + 2, 12), 50)


# ============ 基础数据池 ============
TEAM_SPACES_BASE = [
    (1001, "RED-A-01", "APT追踪组"), (1002, "RED-A-02", "恶意软件分析组"),
    (1003, "RED-B-01", "红蓝对抗组"), (1004, "RED-B-02", "钓鱼演练组"),
    (1005, "RED-C-01", "靶场运营组"), (1006, "RED-C-02", "情报汇聚组"),
]
STAGES = ["UPLOAD", "INDEX", "PARSE", "SEARCH"]
STAGE_NAMES = {"UPLOAD": "上传", "INDEX": "索引", "PARSE": "解析", "SEARCH": "搜索"}
FILE_TYPES = ["pdf", "docx", "eml", "exe", "pcap", "zip", "png", "log", "py", "bin"]
FILE_TYPE_NAMES = {
    "pdf": "PDF文档", "docx": "Word文档", "eml": "邮件文件", "exe": "可执行文件",
    "pcap": "网络抓包", "zip": "压缩包", "png": "图片文件", "log": "日志文件",
    "py": "Python脚本", "bin": "二进制文件",
}
FILE_CATEGORIES = {
    "pdf": "DOCUMENT", "docx": "DOCUMENT", "eml": "DOCUMENT",
    "exe": "OTHER", "pcap": "OTHER", "zip": "ARCHIVE",
    "png": "IMAGE", "log": "DOCUMENT", "py": "CODE", "bin": "OTHER",
}
ERROR_CODES_BASE = {
    "UPLOAD.QUOTA.EXCEED": ("团队空间配额超限", "UPLOAD", "容量", 2),
    "UPLOAD.MIME.REJECT": ("MIME类型被拒", "UPLOAD", "安全", 2),
    "UPLOAD.STORAGE.ERR": ("对象存储异常", "UPLOAD", "基础设施", 1),
    "UPLOAD.NETWORK.ERR": ("上传网络中断", "UPLOAD", "网络", 2),
    "INDEX.ES.REJECTED": ("ES写入拒绝", "INDEX", "基础设施", 1),
    "INDEX.MAPPING.ERR": ("映射错误", "INDEX", "配置", 2),
    "INDEX.TIMEOUT": ("索引超时", "INDEX", "性能", 2),
    "INDEX.ES.OOM": ("ES内存溢出", "INDEX", "基础设施", 1),
    "PARSE.CORRUPT": ("文件损坏", "PARSE", "数据", 2),
    "PARSE.PASSWORD": ("加密文件", "PARSE", "数据", 3),
    "PARSE.OOM": ("内存溢出", "PARSE", "基础设施", 1),
    "PARSE.TIMEOUT": ("解析超时", "PARSE", "性能", 2),
    "PARSE.UNSUPPORTED": ("类型不支持", "PARSE", "配置", 3),
    "SEARCH.ES.TIMEOUT": ("搜索超时", "SEARCH", "性能", 2),
    "SEARCH.QUERY.ERR": ("查询语法错误", "SEARCH", "用户", 3),
    "SEARCH.NO.RESULT": ("无匹配结果", "SEARCH", "数据", 3),
}
OPERATORS = [20011, 20012, 20013, 20014, 20015, 20016, 20017, 20018]
HOT_QUERIES = [
    "apt29", "cobalt strike", "powershell", "mimikatz", "lateral movement",
    "phishing email", "ransomware", "backdoor", "c2 server", "credential dump",
    "lpe exploit", "amsi bypass", "edr evasion", "kernel driver", "shellcode",
    "persistence", "privilege escalation", "data exfiltration", "domain admin", "kerberoasting",
]
SLO_DEFS = [
    (1, "上传可用性", "slo.upload.availability", "UPLOAD", 99.9, "%", "sum(rate(file_upload_success[5m])) / sum(rate(file_upload_total[5m])) * 100", 0.1, 30),
    (2, "索引可搜时延 P95", "slo.index.freshness.p95", "INDEX", 60, "s", "histogram_quantile(0.95, file_index_duration_seconds_bucket)", 5, 30),
    (3, "解析成功率", "slo.parse.success.rate", "PARSE", 95, "%", "sum(file_parse_success) / sum(file_parse_total) * 100", 5, 30),
    (4, "搜索 P95 时延", "slo.search.latency.p95", "SEARCH", 500, "ms", "histogram_quantile(0.95, file_search_duration_ms_bucket)", 0.5, 30),
    (5, "搜索可用性", "slo.search.availability", "SEARCH", 99.5, "%", "sum(rate(file_search_success[5m])) / sum(rate(file_search_total[5m])) * 100", 0.5, 30),
]


def rand_int(a, b):
    return random.randint(a, b)


def rand_float(a, b, nd=2):
    return round(random.uniform(a, b), nd)


def fmt_dt(dt):
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def fmt_date(dt):
    return dt.strftime("%Y-%m-%d")


# ============ 1. 指标字典 t_metric_dict ============
def gen_metric_dict():
    headers = ["指标代码", "指标中文名", "类型", "单位", "所属层", "业务阶段", "口径定义", "允许维度", "启用"]
    # L1-L6 全量指标(基于文档指标体系)
    metrics = [
        # L1 业务流程
        ("file.upload.count", "文件上传总数", "COUNTER", "次", "L1业务", "UPLOAD", "统计周期内上传请求总数(含成功失败)", "team_space_id,file_type,source_type,operator_id", 1),
        ("file.upload.success.count", "上传成功数", "COUNTER", "次", "L1业务", "UPLOAD", "上传成功事件数", "team_space_id,file_type,source_type", 1),
        ("file.upload.success.rate", "上传成功率", "GAUGE", "%", "L1业务", "UPLOAD", "成功数/总数*100", "team_space_id,file_type", 1),
        ("file.upload.duration.p95", "上传耗时P95", "TIMER", "ms", "L1业务", "UPLOAD", "上传耗时95分位", "team_space_id,file_type", 1),
        ("file.upload.bytes", "上传字节数", "COUNTER", "B", "L1业务", "UPLOAD", "上传文件字节累加", "team_space_id,file_type", 1),
        ("file.index.count", "索引建立总数", "COUNTER", "次", "L1业务", "INDEX", "索引建立请求总数", "team_space_id,file_type", 1),
        ("file.index.success.rate", "索引成功率", "GAUGE", "%", "L1业务", "INDEX", "索引成功数/总数*100", "team_space_id", 1),
        ("file.index.freshness", "索引可搜时延", "TIMER", "s", "L1业务", "INDEX", "上传完成到可被搜索的时延", "team_space_id", 1),
        ("file.index.duration.p95", "索引耗时P95", "TIMER", "ms", "L1业务", "INDEX", "索引建立耗时95分位", "team_space_id", 1),
        ("file.parse.count", "解析总数", "COUNTER", "次", "L1业务", "PARSE", "文件解析请求总数", "team_space_id,file_type", 1),
        ("file.parse.success.rate", "解析成功率", "GAUGE", "%", "L1业务", "PARSE", "解析成功数/总数*100", "team_space_id,file_type", 1),
        ("file.parse.duration.p95", "解析耗时P95", "TIMER", "ms", "L1业务", "PARSE", "解析耗时95分位", "team_space_id,file_type", 1),
        ("file.parse.ioc.count", "IOC抽取数", "COUNTER", "个", "L1业务", "PARSE", "解析抽取的IOC总数", "team_space_id,file_type", 1),
        ("file.search.count", "搜索请求总数", "COUNTER", "次", "L1业务", "SEARCH", "搜索请求总数(含所有查询类型)", "team_space_id,query_type", 1),
        ("file.search.success.rate", "搜索成功率", "GAUGE", "%", "L1业务", "SEARCH", "搜索成功数/总数*100", "team_space_id,query_type", 1),
        ("file.search.duration.p95", "搜索耗时P95", "TIMER", "ms", "L1业务", "SEARCH", "搜索耗时95分位", "team_space_id,query_type", 1),
        ("file.search.zero.hit.rate", "零命中率", "GAUGE", "%", "L1业务", "SEARCH", "结果数为0的搜索占比", "team_space_id,query_type", 1),
        ("file.search.result.avg", "平均结果数", "GAUGE", "条", "L1业务", "SEARCH", "搜索返回结果数平均值", "team_space_id,query_type", 1),
        # L2 接口与服务
        ("api.request.qps", "接口QPS", "GAUGE", "req/s", "L2接口", None, "各REST接口每秒请求数", "api_path,method", 1),
        ("api.request.latency.p95", "接口耗时P95", "TIMER", "ms", "L2接口", None, "接口响应耗时95分位", "api_path,method", 1),
        ("api.request.error.rate", "接口错误率", "GAUGE", "%", "L2接口", None, "5xx响应占比", "api_path,method", 1),
        ("api.request.count", "接口请求总数", "COUNTER", "次", "L2接口", None, "接口请求总数", "api_path,method", 1),
        ("service.uptime", "服务可用时长", "GAUGE", "s", "L2接口", None, "服务连续可用秒数", "service_name", 1),
        ("mq.lag", "消息队列积压", "GAUGE", "条", "L2接口", None, "消费端积压消息数", "queue_name,team_space_id", 1),
        ("mq.consume.rate", "消费吞吐", "GAUGE", "msg/s", "L2接口", None, "每秒消费消息数", "queue_name", 1),
        # L3 基础设施
        ("infra.cpu.usage", "CPU使用率", "GAUGE", "%", "L3基础", None, "节点CPU使用率", "node,service_name", 1),
        ("infra.mem.usage", "内存使用率", "GAUGE", "%", "L3基础", None, "节点内存使用率", "node,service_name", 1),
        ("infra.disk.usage", "磁盘使用率", "GAUGE", "%", "L3基础", None, "磁盘使用率", "node,mount_point", 1),
        ("infra.disk.io", "磁盘IO", "GAUGE", "MB/s", "L3基础", None, "磁盘读写吞吐", "node,mount_point", 1),
        ("infra.net.traffic", "网络流量", "GAUGE", "Mbps", "L3基础", None, "网络出入流量", "node,interface,direction", 1),
        ("es.cluster.health", "ES集群健康度", "GAUGE", "枚举", "L3基础", None, "green/yellow/red", "cluster", 1),
        ("es.index.size", "ES索引大小", "GAUGE", "GB", "L3基础", None, "ES索引存储大小", "index_name", 1),
        ("oss.request.rate", "对象存储请求率", "GAUGE", "req/s", "L3基础", None, "OSS每秒请求数", "bucket,operation", 1),
        ("db.connection.active", "DB活跃连接", "GAUGE", "个", "L3基础", None, "数据库活跃连接数", "db_instance", 1),
        ("db.slow.query", "DB慢查询", "COUNTER", "次", "L3基础", None, "慢查询次数(>1s)", "db_instance", 1),
        # L4 安全合规
        ("security.auth.fail.count", "认证失败次数", "COUNTER", "次", "L4安全", None, "登录/鉴权失败次数", "user_id,source_ip", 1),
        ("security.dangerous.file", "危险文件数", "COUNTER", "个", "L4安全", "UPLOAD", "检出为恶意的文件数", "team_space_id,detect_engine", 1),
        ("security.sensitive.access", "敏感数据访问", "COUNTER", "次", "L4安全", None, "敏感文件访问次数", "team_space_id,user_id", 1),
        ("security.abnormal.download", "异常下载", "COUNTER", "次", "L4安全", None, "批量/异常下载行为次数", "team_space_id,user_id", 1),
        ("security.sql.inject", "SQL注入尝试", "COUNTER", "次", "L4安全", "SEARCH", "搜索词中SQL注入特征次数", "team_space_id", 1),
        ("audit.log.integrity", "审计日志完整性", "GAUGE", "%", "L4安全", None, "审计日志未丢失占比", "log_type", 1),
        # L5 SLO
        ("slo.upload.availability", "上传可用性SLO", "GAUGE", "%", "L5SLO", "UPLOAD", "30天滚动上传成功率", "team_space_id", 1),
        ("slo.index.freshness.p95", "索引可搜时延SLO", "GAUGE", "s", "L5SLO", "INDEX", "30天P95索引可搜时延", "team_space_id", 1),
        ("slo.parse.success.rate", "解析成功率SLO", "GAUGE", "%", "L5SLO", "PARSE", "30天滚动解析成功率", "team_space_id", 1),
        ("slo.search.latency.p95", "搜索时延SLO", "GAUGE", "ms", "L5SLO", "SEARCH", "30天P95搜索时延", "team_space_id", 1),
        ("slo.search.availability", "搜索可用性SLO", "GAUGE", "%", "L5SLO", "SEARCH", "30天滚动搜索成功率", "team_space_id", 1),
        ("slo.error.budget.remaining", "剩余错误预算", "GAUGE", "%", "L5SLO", None, "SLO错误预算剩余比例", "slo_code,team_space_id", 1),
        ("slo.burn.rate.2h", "2小时燃烧率", "GAUGE", "倍", "L5SLO", None, "2小时错误消耗速率", "slo_code,team_space_id", 1),
        ("slo.burn.rate.6h", "6小时燃烧率", "GAUGE", "倍", "L5SLO", None, "6小时错误消耗速率", "slo_code,team_space_id", 1),
        # L6 数据质量与容量成本
        ("quality.duplicate.rate", "重复文件率", "GAUGE", "%", "L6质量", "UPLOAD", "MD5重复文件占比", "team_space_id", 1),
        ("quality.parse.fidelity", "解析保真度", "GAUGE", "%", "L6质量", "PARSE", "解析结果与原文一致率", "team_space_id,file_type", 1),
        ("quality.empty.file.rate", "空文件率", "GAUGE", "%", "L6质量", "UPLOAD", "0字节文件占比", "team_space_id", 1),
        ("quality.tag.missing.rate", "标签缺失率", "GAUGE", "%", "L6质量", "INDEX", "缺少必填标签的文件占比", "team_space_id", 1),
        ("capacity.storage.used", "存储用量", "GAUGE", "B", "L6容量", None, "团队空间已用存储", "team_space_id", 1),
        ("capacity.storage.quota", "存储配额", "GAUGE", "B", "L6容量", None, "团队空间存储配额", "team_space_id", 1),
        ("capacity.storage.usage.rate", "存储使用率", "GAUGE", "%", "L6容量", None, "已用/配额*100", "team_space_id", 1),
        ("capacity.file.count", "文件总数", "GAUGE", "个", "L6容量", None, "团队空间文件数", "team_space_id", 1),
        ("capacity.cost.monthly", "月度成本", "GAUGE", "元", "L6容量", None, "月度存储+计算成本", "team_space_id,cost_type", 1),
        ("capacity.growth.rate", "存储增长率", "GAUGE", "%", "L6容量", None, "日存储增长率", "team_space_id", 1),
    ]
    # 不足100行,补充细分维度变体
    variants = [
        ("file.upload.count.by.source", "上传数(按来源)", "COUNTER", "次", "L1业务", "UPLOAD"),
        ("file.parse.count.by.type", "解析数(按类型)", "COUNTER", "次", "L1业务", "PARSE"),
        ("file.search.count.by.type", "搜索数(按类型)", "COUNTER", "次", "L1业务", "SEARCH"),
    ]
    extra_types = ["pdf", "docx", "eml", "exe", "pcap", "zip", "png", "log", "py", "bin"]
    extra_spaces = ["ts_1001", "ts_1002", "ts_1003", "ts_1004", "ts_1005", "ts_1006"]
    idx = 0
    while len(metrics) < ROWS_PER_SHEET:
        base = variants[idx % len(variants)]
        ft = extra_types[idx % len(extra_types)]
        ts = extra_spaces[idx % len(extra_spaces)]
        metrics.append((
            f"{base[0]}.{ft}.{ts}", f"{base[1]}-{ft}-{ts}", base[2], base[3], base[4], base[5],
            f"{base[1]}按文件类型{ft}与团队空间{ts}细分", "team_space_id,file_type", 1
        ))
        idx += 1
    rows = [m[:9] for m in metrics[:ROWS_PER_SHEET]]
    return headers, rows


# ============ 2. 团队空间 t_team_space ============
def gen_team_space():
    headers = ["团队空间ID", "租户ID", "空间编码", "空间名称", "存储配额(字节)", "文件数配额", "状态", "空间负责人ID", "创建时间", "更新时间"]
    rows = []
    base_names = ["APT追踪", "恶意软件分析", "红蓝对抗", "钓鱼演练", "靶场运营", "情报汇聚",
                  "样本归档", "漏洞研究", "应急响应", "威胁狩猎"]
    for i in range(ROWS_PER_SHEET):
        ts_id = 1001 + i
        tenant_id = 1
        code = f"RED-{chr(65 + i % 6)}-{(i // 6 + 1):02d}"
        name = f"{base_names[i % len(base_names)]}组{i // len(base_names) + 1:02d}"
        quota = random.choice([200, 300, 500, 800, 1024, 2048]) * 1024 ** 3
        file_quota = random.choice([20000, 30000, 50000, 80000, 100000, 200000])
        status = 1 if random.random() > 0.05 else 0
        owner = 20011 + (i % 8)
        created = datetime(2025, 1, 1) + timedelta(days=i * 3, hours=i % 24)
        updated = created + timedelta(days=random.randint(1, 30))
        rows.append([ts_id, tenant_id, code, name, quota, file_quota, status, owner, fmt_dt(created), fmt_dt(updated)])
    return headers, rows


# ============ 3. 文件事件流 t_file_event ============
def gen_file_event():
    headers = ["事件ID", "链路追踪ID", "团队空间ID", "文件ID", "业务阶段", "事件类型", "耗时(毫秒)",
               "文件大小(字节)", "文件类型", "来源类型", "操作人ID", "错误码", "错误信息", "查询类型", "查询关键词", "结果数", "IOC抽取数", "事件时间"]
    rows = []
    now = datetime(2026, 7, 27, 10, 0, 0)
    err_codes_by_stage = {
        "UPLOAD": ["UPLOAD.QUOTA.EXCEED", "UPLOAD.MIME.REJECT", "UPLOAD.STORAGE.ERR", "UPLOAD.NETWORK.ERR"],
        "INDEX": ["INDEX.ES.REJECTED", "INDEX.MAPPING.ERR", "INDEX.TIMEOUT", "INDEX.ES.OOM"],
        "PARSE": ["PARSE.CORRUPT", "PARSE.PASSWORD", "PARSE.OOM", "PARSE.TIMEOUT", "PARSE.UNSUPPORTED"],
        "SEARCH": ["SEARCH.ES.TIMEOUT", "SEARCH.QUERY.ERR", "SEARCH.NO.RESULT"],
    }
    err_msgs = {
        "UPLOAD.QUOTA.EXCEED": "团队空间存储配额已超限,请扩容或清理",
        "UPLOAD.MIME.REJECT": "MIME类型 application/x-msdownload 不在白名单",
        "UPLOAD.STORAGE.ERR": "对象存储写入失败:Connection reset",
        "UPLOAD.NETWORK.ERR": "客户端上传中断,断点续传失败",
        "INDEX.ES.REJECTED": "ES写入拒绝(429):bulk queue filled",
        "INDEX.MAPPING.ERR": "字段file_size映射冲突:long vs keyword",
        "INDEX.TIMEOUT": "ES bulk写入超时(30s)",
        "INDEX.ES.OOM": "ES节点Old Gen使用率>90%,触发熔断",
        "PARSE.CORRUPT": "文件头损坏,无法识别格式",
        "PARSE.PASSWORD": "文件已加密,未提供密码",
        "PARSE.OOM": "解析进程OOM,文件过大(>500MB)",
        "PARSE.TIMEOUT": "解析超时(60s),文件结构复杂",
        "PARSE.UNSUPPORTED": "暂不支持的文件类型:iso",
        "SEARCH.ES.TIMEOUT": "ES查询超时(10s),结果集过大",
        "SEARCH.QUERY.ERR": "查询语法错误:未闭合的引号",
        "SEARCH.NO.RESULT": "无匹配结果,建议扩大查询范围",
    }
    for i in range(ROWS_PER_SHEET):
        eid = 100000000 + i
        trace_id = f"trace-{random.randint(0x10000000, 0xffffffff):x}"
        ts_id = 1001 + (i % 6)
        file_id = random.randint(100000, 999999)
        stage = STAGES[i % 4]
        is_fail = random.random() < 0.18
        event_type = "FAIL" if is_fail else "SUCCESS"
        if stage == "PARSE":
            duration = rand_int(1500, 9000)
        elif stage == "UPLOAD":
            duration = rand_int(300, 2000)
        else:
            duration = rand_int(100, 800)
        file_size = rand_int(1024, 200 * 1024 * 1024)
        ft = FILE_TYPES[i % len(FILE_TYPES)]
        source = random.choice([1, 1, 2, 3])  # 上传为主
        operator = OPERATORS[i % len(OPERATORS)]
        err_code = random.choice(err_codes_by_stage[stage]) if is_fail else None
        err_msg = err_msgs.get(err_code, "") if is_fail else None
        query_type = random.choice(["KEYWORD", "SEMANTIC", "FUZZY"]) if stage == "SEARCH" else None
        query_kw = random.choice(HOT_QUERIES) if stage == "SEARCH" else None
        result_cnt = rand_int(0, 500) if stage == "SEARCH" else None
        ioc_cnt = rand_int(0, 20) if stage == "PARSE" and not is_fail else None
        created = now - timedelta(seconds=i * 45 + rand_int(5, 40))
        rows.append([eid, trace_id, ts_id, file_id, stage, event_type, duration, file_size, ft,
                     source, operator, err_code, err_msg, query_type, query_kw, result_cnt, ioc_cnt, fmt_dt(created)])
    return headers, rows


# ============ 4. 小时聚合指标 t_metric_hourly ============
def gen_metric_hourly():
    headers = ["聚合ID", "统计小时", "团队空间ID", "业务阶段", "指标代码", "总数", "成功数", "失败数",
               "字节数累加", "耗时P50(ms)", "耗时P95(ms)", "耗时P99(ms)", "平均耗时(ms)", "失败Top错误码", "文件类型维度", "查询类型维度"]
    metric_codes = {
        "UPLOAD": ["file.upload.count", "file.upload.duration.p95", "file.upload.bytes"],
        "INDEX": ["file.index.count", "file.index.duration.p95"],
        "PARSE": ["file.parse.count", "file.parse.duration.p95", "file.parse.ioc.count"],
        "SEARCH": ["file.search.count", "file.search.duration.p95", "file.search.result.avg"],
    }
    err_by_stage = {
        "UPLOAD": "UPLOAD.QUOTA.EXCEED", "INDEX": "INDEX.ES.REJECTED",
        "PARSE": "PARSE.TIMEOUT", "SEARCH": "SEARCH.ES.TIMEOUT",
    }
    base_p95 = {"UPLOAD": 1200, "INDEX": 400, "PARSE": 5000, "SEARCH": 350}
    rows = []
    now = datetime(2026, 7, 27, 9, 0, 0)
    rid = 1
    for i in range(ROWS_PER_SHEET):
        stat_time = now - timedelta(hours=i)
        ts_id = 1001 + (i % 6)
        stage = STAGES[i % 4]
        mc = random.choice(metric_codes[stage])
        total = rand_int(100, 800)
        fail = rand_int(0, max(2, total // 20))
        succ = total - fail
        bytes_total = rand_int(100 * 1024 * 1024, 50 * 1024 ** 3) if stage in ("UPLOAD", "PARSE") else 0
        p50 = base_p95[stage] - rand_int(100, 400)
        p95 = base_p95[stage] + rand_int(-100, 300)
        p99 = p95 + rand_int(200, 800)
        avg = p50 - rand_int(50, 200)
        fail_code = err_by_stage[stage] if fail > 0 else None
        dim_ft = random.choice(FILE_TYPES) if stage in ("UPLOAD", "PARSE") else None
        dim_qt = random.choice(["KEYWORD", "SEMANTIC", "FUZZY"]) if stage == "SEARCH" else None
        rows.append([rid, fmt_dt(stat_time), ts_id, stage, mc, total, succ, fail, bytes_total,
                     p50, p95, p99, avg, fail_code, dim_ft, dim_qt])
        rid += 1
    return headers, rows


# ============ 5. 日聚合指标 t_metric_daily ============
def gen_metric_daily():
    headers = ["聚合ID", "统计日期", "团队空间ID", "业务阶段", "指标代码", "总数", "成功数", "失败数",
               "字节数累加", "耗时P50(ms)", "耗时P95(ms)", "耗时P99(ms)", "成功率(%)", "失败Top错误码"]
    metric_codes = {
        "UPLOAD": ["file.upload.count", "file.upload.duration.p95"],
        "INDEX": ["file.index.count", "file.index.freshness"],
        "PARSE": ["file.parse.count", "file.parse.duration.p95"],
        "SEARCH": ["file.search.count", "file.search.duration.p95"],
    }
    err_by_stage = {
        "UPLOAD": "UPLOAD.QUOTA.EXCEED", "INDEX": "INDEX.TIMEOUT",
        "PARSE": "PARSE.TIMEOUT", "SEARCH": "SEARCH.ES.TIMEOUT",
    }
    base_p95 = {"UPLOAD": 1100, "INDEX": 350, "PARSE": 4500, "SEARCH": 300}
    rows = []
    rid = 1
    today = datetime(2026, 7, 27).date()
    for i in range(ROWS_PER_SHEET):
        stat_date = today - timedelta(days=i // 6)
        ts_id = 1001 + (i % 6)
        stage = STAGES[i % 4]
        mc = random.choice(metric_codes[stage])
        total = rand_int(2000, 30000)
        fail = rand_int(10, max(50, total // 15))
        succ = total - fail
        rate = round(succ / total * 100, 2)
        bytes_total = rand_int(1 * 1024 ** 3, 200 * 1024 ** 3) if stage == "UPLOAD" else 0
        p50 = base_p95[stage] - rand_int(50, 300)
        p95 = base_p95[stage] + rand_int(-50, 200)
        p99 = p95 + rand_int(150, 600)
        fail_code = err_by_stage[stage]
        rows.append([rid, fmt_date(datetime.combine(stat_date, datetime.min.time())), ts_id, stage, mc,
                     total, succ, fail, bytes_total, p50, p95, p99, rate, fail_code])
        rid += 1
    return headers, rows


# ============ 6. TopN记录 t_topn_record ============
def gen_topn_record():
    headers = ["记录ID", "统计日期", "团队空间ID", "TopN类型", "排名", "项标识", "出现次数", "附加信息"]
    topn_types = [
        ("HOT_QUERY", HOT_QUERIES, (50, 800)),
        ("ZERO_HIT_QUERY", [f"{q} v2" for q in HOT_QUERIES], (10, 200)),
        ("FAIL_FILE", FILE_TYPES, (20, 400)),
        ("LARGE_FILE", FILE_TYPES, (5, 100)),
    ]
    rows = []
    rid = 1
    today = datetime(2026, 7, 27).date()
    for i in range(ROWS_PER_SHEET):
        stat_date = today - timedelta(days=i // 20)
        ts_id = 1001 + (i % 6)
        t_type, items, cnt_range = topn_types[i % len(topn_types)]
        rank = (i // len(topn_types)) % 20 + 1
        item = items[i % len(items)]
        cnt = rand_int(cnt_range[0], cnt_range[1])
        extra = f'{{"rank":{rank},"trend":{"up" if random.random() > 0.4 else "down"}}}'
        rows.append([rid, fmt_date(datetime.combine(stat_date, datetime.min.time())), ts_id, t_type, rank, item, cnt, extra])
        rid += 1
    return headers, rows


# ============ 7. 错误码字典 t_dim_error_code ============
def gen_dim_error_code():
    headers = ["错误码", "错误名称", "所属阶段", "错误类别", "严重级别", "处理建议", "启用"]
    suggestions = {
        "UPLOAD.QUOTA.EXCEED": "扩容团队空间配额或清理历史文件",
        "UPLOAD.MIME.REJECT": "检查文件类型是否在白名单,必要时调整MIME策略",
        "UPLOAD.STORAGE.ERR": "检查对象存储服务状态与网络,重试上传",
        "UPLOAD.NETWORK.ERR": "检查客户端网络,启用断点续传",
        "INDEX.ES.REJECTED": "降低bulk写入并发,扩容ES集群",
        "INDEX.MAPPING.ERR": "修正索引映射,清理冲突字段",
        "INDEX.TIMEOUT": "优化bulk大小,检查ES负载",
        "INDEX.ES.OOM": "扩容ES内存,调整JVM Old Gen回收策略",
        "PARSE.CORRUPT": "重新上传源文件,校验MD5",
        "PARSE.PASSWORD": "提供文件解密密码",
        "PARSE.OOM": "限制单文件大小,分片解析大文件",
        "PARSE.TIMEOUT": "优化解析器,增大超时阈值",
        "PARSE.UNSUPPORTED": "扩展解析器支持的文件类型",
        "SEARCH.ES.TIMEOUT": "优化查询语句,限制结果集大小",
        "SEARCH.QUERY.ERR": "检查查询语法,使用查询构建器",
        "SEARCH.NO.RESULT": "扩大查询范围或使用模糊查询",
    }
    rows = []
    # 基础错误码
    for code, (name, stage, cat, sev) in ERROR_CODES_BASE.items():
        rows.append([code, name, stage, cat, sev, suggestions.get(code, ""), 1])
    # 扩展细分错误码到100行
    ext_stages = ["UPLOAD", "INDEX", "PARSE", "SEARCH"]
    ext_categories = ["容量", "安全", "性能", "配置", "网络", "数据", "用户", "基础设施"]
    ext_names_pool = ["校验失败", "格式异常", "版本不兼容", "权限不足", "连接中断", "数据缺失",
                      "配置错误", "限流触发", "签名无效", "编码错误", "字段缺失", "重复提交",
                      "超限拒绝", "服务降级", "资源不足", "依赖异常", "调度失败", "校验超时"]
    idx = 0
    while len(rows) < ROWS_PER_SHEET:
        stage = ext_stages[idx % 4]
        cat = ext_categories[idx % len(ext_categories)]
        sev = (idx % 3) + 1
        suffix = f"{cat.upper().replace('基础设施', 'INFRA')}{idx:03d}"
        code = f"{stage}.{suffix}"
        name = ext_names_pool[idx % len(ext_names_pool)] + f"({stage})"
        sug = "请查看日志并联系平台运维"
        rows.append([code, name, stage, cat, sev, sug, 1])
        idx += 1
    return headers, rows[:ROWS_PER_SHEET]


# ============ 8. 文件类型字典 t_dim_file_type ============
def gen_dim_file_type():
    headers = ["文件类型", "类型中文名", "分类", "是否支持解析", "图标标识"]
    rows = []
    # 基础类型
    for ft in FILE_TYPES:
        rows.append([ft, FILE_TYPE_NAMES[ft], FILE_CATEGORIES[ft], 1, f"icon-{ft}"])
    # 扩展到100行(细分类型/版本)
    ext_pool = [
        ("doc", "Word旧版文档", "DOCUMENT"), ("xls", "Excel旧版表格", "DOCUMENT"), ("ppt", "PPT演示", "DOCUMENT"),
        ("xlsx", "Excel表格", "DOCUMENT"), ("pptx", "PPT演示", "DOCUMENT"), ("txt", "文本文件", "DOCUMENT"),
        ("md", "Markdown文档", "DOCUMENT"), ("json", "JSON数据", "DOCUMENT"), ("xml", "XML文件", "DOCUMENT"),
        ("csv", "CSV表格", "DOCUMENT"), ("html", "HTML网页", "DOCUMENT"), ("js", "JavaScript脚本", "CODE"),
        ("ts", "TypeScript脚本", "CODE"), ("java", "Java源码", "CODE"), ("c", "C源码", "CODE"),
        ("cpp", "C++源码", "CODE"), ("go", "Go源码", "CODE"), ("rb", "Ruby源码", "CODE"),
        ("php", "PHP源码", "CODE"), ("sh", "Shell脚本", "CODE"), ("bat", "批处理", "CODE"),
        ("ps1", "PowerShell脚本", "CODE"), ("vbs", "VBScript", "CODE"), ("jar", "Java归档", "CODE"),
        ("class", "Java类文件", "CODE"), ("dll", "动态链接库", "OTHER"), ("so", "Linux共享库", "OTHER"),
        ("sys", "系统驱动", "OTHER"), ("img", "磁盘镜像", "ARCHIVE"), ("iso", "ISO镜像", "ARCHIVE"),
        ("rar", "RAR压缩包", "ARCHIVE"), ("7z", "7Z压缩包", "ARCHIVE"), ("gz", "Gzip压缩", "ARCHIVE"),
        ("tar", "TAR归档", "ARCHIVE"), ("jpg", "JPEG图片", "IMAGE"), ("jpeg", "JPEG图片", "IMAGE"),
        ("gif", "GIF图片", "IMAGE"), ("bmp", "位图", "IMAGE"), ("webp", "WebP图片", "IMAGE"),
        ("svg", "SVG矢量图", "IMAGE"), ("tif", "TIFF图片", "IMAGE"), ("mp4", "MP4视频", "VIDEO"),
        ("avi", "AVI视频", "VIDEO"), ("mov", "MOV视频", "VIDEO"), ("wmv", "WMV视频", "VIDEO"),
        ("mkv", "MKV视频", "VIDEO"), ("mp3", "MP3音频", "AUDIO"), ("wav", "WAV音频", "AUDIO"),
        ("flac", "FLAC音频", "AUDIO"), ("aac", "AAC音频", "AUDIO"), ("ogg", "OGG音频", "AUDIO"),
        ("m4a", "M4A音频", "AUDIO"), ("msg", "Outlook邮件", "DOCUMENT"), ("oft", "Outlook模板", "DOCUMENT"),
        ("one", "OneNote", "DOCUMENT"), ("rtf", "RTF富文本", "DOCUMENT"), ("wps", "WPS文档", "DOCUMENT"),
        ("et", "WPS表格", "DOCUMENT"), ("dps", "WPS演示", "DOCUMENT"), ("sql", "SQL脚本", "CODE"),
        ("yml", "YAML配置", "CODE"), ("yaml", "YAML配置", "CODE"), ("toml", "TOML配置", "CODE"),
        ("ini", "INI配置", "CODE"), ("conf", "配置文件", "CODE"), ("reg", "注册表", "CODE"),
        ("hiv", "注册表蜂巢", "OTHER"), ("evtx", "Windows事件日志", "DOCUMENT"), ("dmp", "内存转储", "OTHER"),
        ("raw", "原始镜像", "ARCHIVE"), ("vmdk", "VMware磁盘", "ARCHIVE"), ("vhd", "VHD磁盘", "ARCHIVE"),
        ("vhdx", "VHDX磁盘", "ARCHIVE"), ("ova", "OVA虚拟设备", "ARCHIVE"), ("ovf", "OVF描述", "DOCUMENT"),
        ("pcapng", "PCAP下一代抓包", "OTHER"), ("cap", "CAP抓包", "OTHER"), ("key", "密钥文件", "OTHER"),
        ("pem", "PEM证书", "OTHER"), ("crt", "证书文件", "OTHER"), ("p12", "PKCS12证书", "OTHER"),
        ("kubeconfig", "K8s配置", "CODE"), ("dockerfile", "Dockerfile", "CODE"),
    ]
    idx = 0
    while len(rows) < ROWS_PER_SHEET:
        if idx < len(ext_pool):
            ft_ext, name, cat = ext_pool[idx]
            parse_supported = 1 if cat in ("DOCUMENT", "CODE", "ARCHIVE") else (0 if idx % 7 == 6 else 1)
            rows.append([ft_ext, name, cat, parse_supported, f"icon-{ft_ext}"])
        else:
            ft_ext = f"type{idx:03d}"
            cat = ["DOCUMENT", "IMAGE", "VIDEO", "AUDIO", "ARCHIVE", "CODE", "OTHER"][idx % 7]
            rows.append([ft_ext, f"自定义类型{idx:03d}", cat, 1 if idx % 5 else 0, f"icon-{ft_ext}"])
        idx += 1
    return headers, rows[:ROWS_PER_SHEET]


# ============ 9. SLO定义 t_slo_definition ============
def gen_slo_definition():
    headers = ["SLO ID", "SLO名称", "SLO编码", "业务阶段", "目标值", "目标单位", "计算表达式", "错误预算(%)", "滚动窗口(天)", "启用"]
    rows = []
    # 基础5个SLO
    for s in SLO_DEFS:
        rows.append(list(s) + [1])
    # 扩展到100行:按团队空间×阶段细分
    idx = 0
    while len(rows) < ROWS_PER_SHEET:
        base = SLO_DEFS[idx % len(SLO_DEFS)]
        ts_seq = idx // len(SLO_DEFS) + 1
        ts_id = 1000 + ts_seq
        rows.append([
            len(rows) + 1,
            f"{base[1]}-空间{ts_id}",
            f"{base[2]}.ts{ts_id}",
            base[3],
            base[4],
            base[5],
            base[6],
            base[7],
            base[8],
            1,
        ])
        idx += 1
    return headers, rows[:ROWS_PER_SHEET]


# ============ 10. SLO记录 t_slo_record ============
def gen_slo_record():
    headers = ["记录ID", "统计日期", "SLO ID", "团队空间ID", "实际值", "剩余错误预算(%)", "2小时燃烧率", "6小时燃烧率", "状态"]
    # SLO基础参数(对应5个SLO定义)
    slo_params = {
        1: {"actual": (99.75, 99.99), "budget": (40, 95), "burn2h": (0.1, 1.5), "burn6h": (0.1, 1.2)},
        2: {"actual": (35, 75),       "budget": (30, 90), "burn2h": (0.2, 2.0), "burn6h": (0.2, 1.6)},
        3: {"actual": (92, 97.5),     "budget": (20, 85), "burn2h": (0.5, 3.0), "burn6h": (0.4, 2.5)},
        4: {"actual": (380, 620),     "budget": (35, 92), "burn2h": (0.1, 1.8), "burn6h": (0.1, 1.4)},
        5: {"actual": (99.3, 99.95),  "budget": (50, 95), "burn2h": (0.1, 1.0), "burn6h": (0.1, 0.8)},
    }
    rows = []
    rid = 1
    today = datetime(2026, 7, 27).date()
    for i in range(ROWS_PER_SHEET):
        stat_date = today - timedelta(days=i // 5)
        slo_id = (i % 5) + 1
        ts_id = 1001 + (i % 6)
        p = slo_params[slo_id]
        actual = rand_float(*p["actual"])
        budget = rand_float(*p["budget"])
        burn2h = rand_float(*p["burn2h"])
        burn6h = rand_float(*p["burn6h"])
        # 状态:0正常 1告警 2违约
        target = {1: 99.9, 2: 60, 3: 95, 4: 500, 5: 99.5}[slo_id]
        is_good = (slo_id in (1, 3, 5) and actual >= target) or (slo_id in (2, 4) and actual <= target)
        if is_good and budget > 30:
            status = 0
        elif is_good:
            status = 1
        else:
            status = 2
        rows.append([rid, fmt_date(datetime.combine(stat_date, datetime.min.time())), slo_id, ts_id, actual, budget, burn2h, burn6h, status])
        rid += 1
    return headers, rows


# ============ 11. 告警规则 t_alert_rule ============
def gen_alert_rule():
    headers = ["规则ID", "规则名称", "规则编码", "团队空间ID", "业务阶段", "指标表达式", "触发条件", "时间窗口(分钟)", "严重级别", "通知渠道", "启用", "创建时间"]
    rule_templates = [
        ("上传成功率告警", "alert.upload.success", "UPLOAD", "file_upload_success_rate", "< 0.99", 5, 1),
        ("上传P95耗时告警", "alert.upload.latency", "UPLOAD", "file_upload_duration_p95", "> 1500", 5, 2),
        ("索引成功率告警", "alert.index.success", "INDEX", "file_index_success_rate", "< 0.98", 5, 1),
        ("索引可搜时延告警", "alert.index.freshness", "INDEX", "file_index_freshness_seconds", "> 60", 10, 2),
        ("解析成功率告警", "alert.parse.success", "PARSE", "file_parse_success_rate", "< 0.95", 5, 1),
        ("解析P95耗时告警", "alert.parse.latency", "PARSE", "file_parse_duration_p95", "> 8000", 10, 2),
        ("搜索P95耗时告警", "alert.search.latency", "SEARCH", "file_search_duration_p95", "> 500", 5, 2),
        ("零命中率告警", "alert.search.zero", "SEARCH", "file_search_zero_hit_rate", "> 0.15", 15, 3),
        ("存储使用率告警", "alert.storage.usage", None, "capacity_storage_usage_rate", "> 0.85", 60, 2),
        ("队列积压告警", "alert.mq.lag", None, "mq_lag", "> 500", 5, 2),
        ("ES集群红色告警", "alert.es.red", None, "es_cluster_health", "== red", 1, 1),
        ("CPU使用率告警", "alert.cpu.high", None, "infra_cpu_usage", "> 0.85", 5, 2),
        ("认证失败激增告警", "alert.auth.fail", None, "security_auth_fail_count", "> 20", 5, 1),
        ("危险文件告警", "alert.danger.file", "UPLOAD", "security_dangerous_file", "> 0", 1, 1),
    ]
    channels = ["feishu-webhook", "feishu-webhook,sms", "feishu-webhook,email", "email"]
    rows = []
    for i in range(ROWS_PER_SHEET):
        tpl = rule_templates[i % len(rule_templates)]
        ts_seq = i // len(rule_templates)
        ts_id = 1000 + ts_seq + 1 if ts_seq < 6 else None  # 前6个空间各一套,后面全局
        name = f"{tpl[0]}-空间{ts_seq + 1}" if ts_id else tpl[0]
        code = f"{tpl[1]}.ts{ts_seq + 1}" if ts_id else tpl[1]
        expr = tpl[3]
        cond = tpl[4]
        window = tpl[5]
        sev = tpl[6]
        ch = channels[i % len(channels)]
        enabled = 1 if random.random() > 0.1 else 0
        created = datetime(2026, 1, 1) + timedelta(days=i, hours=i % 24)
        rows.append([i + 1, name, code, ts_id, tpl[2], expr, cond, window, sev, ch, enabled, fmt_dt(created)])
    return headers, rows[:ROWS_PER_SHEET]


# ============ 12. 告警事件 t_alert_event ============
def gen_alert_event():
    headers = ["事件ID", "规则ID", "团队空间ID", "严重级别", "触发时间", "恢复时间", "触发时实际值", "上下文", "状态", "通知状态"]
    rule_count = 100  # 假设有100条规则
    trigger_value_tpl = [
        ("99.85%", "成功率0.9985"), ("1850ms", "P95耗时1850ms"), ("97.2%", "成功率0.972"),
        ("78s", "时延78秒"), ("93.1%", "成功率0.931"), ("9200ms", "P95耗时9200ms"),
        ("680ms", "P95耗时680ms"), ("18.5%", "零命中率0.185"), ("88.2%", "存储使用率0.882"),
        ("650", "积压650条"), ("red", "ES集群红色"), ("91%", "CPU使用率0.91"),
        ("25", "认证失败25次"), ("3", "危险文件3个"),
    ]
    rows = []
    now = datetime(2026, 7, 27, 10, 0, 0)
    for i in range(ROWS_PER_SHEET):
        eid = 900000001 + i
        rule_id = (i % rule_count) + 1
        ts_id = 1001 + (i % 6)
        sev = (i % 3) + 1
        fired = now - timedelta(hours=i * 2 + rand_int(0, 60), minutes=rand_int(0, 59))
        # 60%已恢复,30%触发中,10%已忽略
        r = random.random()
        if r < 0.6:
            status = 1
            resolved = fired + timedelta(minutes=rand_int(5, 180))
            resolved_str = fmt_dt(resolved)
        elif r < 0.9:
            status = 0
            resolved_str = None
        else:
            status = 2
            resolved_str = None
        tv_pair = trigger_value_tpl[i % len(trigger_value_tpl)]
        ctx = f'{{"rule_id":{rule_id},"sample":"trace-{random.randint(0x100000,0xffffff):x}","value":"{tv_pair[1]}"}}'
        notify = 1 if random.random() > 0.15 else (2 if random.random() > 0.5 else 0)
        rows.append([eid, rule_id, ts_id, sev, fmt_dt(fired), resolved_str, tv_pair[0], ctx, status, notify])
    return headers, rows


# ============ 主流程 ============
def main():
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认sheet

    sheets = [
        ("1-指标字典", gen_metric_dict),
        ("2-团队空间", gen_team_space),
        ("3-文件事件流", gen_file_event),
        ("4-小时聚合指标", gen_metric_hourly),
        ("5-日聚合指标", gen_metric_daily),
        ("6-TopN记录", gen_topn_record),
        ("7-错误码字典", gen_dim_error_code),
        ("8-文件类型字典", gen_dim_file_type),
        ("9-SLO定义", gen_slo_definition),
        ("10-SLO记录", gen_slo_record),
        ("11-告警规则", gen_alert_rule),
        ("12-告警事件", gen_alert_event),
    ]

    for name, gen_fn in sheets:
        headers, rows = gen_fn()
        ws = wb.create_sheet(title=name)
        write_sheet(ws, headers, rows)
        print(f"  [{name}] 生成 {len(rows)} 行 x {len(headers)} 列")

    wb.save(OUT_PATH)
    print(f"\n✓ 已生成: {OUT_PATH}")
    print(f"  共 {len(sheets)} 个 sheet, 每个 {ROWS_PER_SHEET} 行")


if __name__ == "__main__":
    main()
